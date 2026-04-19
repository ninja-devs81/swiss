import csv
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from typing import Optional
from models.schemas import ReconciliationSummary, PlanActualSummary
from models.enums import ReconciliationStatus, STATUS_LABELS, PlanActualStatus, PLAN_ACTUAL_LABELS, OverallStatus

STATUS_FILL = {
    ReconciliationStatus.WITHIN_LIMIT.value:    "C6EFCE",
    ReconciliationStatus.NEAR_LIMIT.value:      "FFEB9C",
    ReconciliationStatus.EXCEEDED.value:        "FFC7CE",
    ReconciliationStatus.OUT_OF_RANGE.value:    "E8D5FF",
    ReconciliationStatus.NO_SERVICES.value:     "DDDDDD",
    ReconciliationStatus.NO_PRESCRIPTION.value: "FCE4D6",
}
STATUS_FONT_COLOR = {
    ReconciliationStatus.WITHIN_LIMIT.value:    "276221",
    ReconciliationStatus.NEAR_LIMIT.value:      "9C5700",
    ReconciliationStatus.EXCEEDED.value:        "9C0006",
    ReconciliationStatus.OUT_OF_RANGE.value:    "5B21B6",
    ReconciliationStatus.NO_SERVICES.value:     "595959",
    ReconciliationStatus.NO_PRESCRIPTION.value: "843C0C",
}

def _thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def generate_report(
    summary: ReconciliationSummary,
    plan_actual: Optional[PlanActualSummary] = None,
) -> bytes:
    wb = Workbook()

    # ---- Sheet 1: Summary ----
    ws_sum = wb.active
    ws_sum.title = "Übersicht"

    header_font = Font(name="Calibri", bold=True, size=11)
    title_font  = Font(name="Calibri", bold=True, size=14)
    hdr_fill    = PatternFill("solid", fgColor="1F3864")

    ws_sum["A1"] = "Spitex Leistungskontrolle – Reconciliation Report"
    ws_sum["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F3864")
    ws_sum["A2"] = f"Erstellt am: {summary.generated_at}"
    ws_sum["A2"].font = Font(name="Calibri", size=10, color="595959")
    ws_sum.merge_cells("A1:D1")
    ws_sum.merge_cells("A2:D2")

    ws_sum.append([])
    ws_sum.append(["Kennzahl", "Anzahl"])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = hdr_fill

    stats = [
        ("Patienten gesamt",       summary.total_patients),
        ("Verordnungen",           summary.total_prescriptions),
        ("Im Rahmen",              summary.within_limit),
        ("Nahe am Limit (≥ 80%)", summary.near_limit),
        ("Überschritten",          summary.exceeded),
        ("Ausserhalb Zeitraum",    summary.out_of_range),
        ("Keine Leistungen",       summary.no_services),
        ("Keine Verordnung",       summary.no_prescription),
    ]
    for label, val in stats:
        ws_sum.append([label, val])

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15

    # ---- Sheet 2: Detail ----
    ws_det = wb.create_sheet("Detailansicht")

    headers = [
        "Patient", "Geburtsdatum", "Tarifcode", "Tarifziffer",
        "Gültig von", "Gültig bis",
        "Bewilligte Min.", "Erbrachte Min.", "Verbleibend Min.",
        "Auslastung %", "Leistungen (Anz.)", "Ausserhalb (Anz.)", "Status", "Ablehnungsgrund",
    ]
    ws_det.append(headers)
    for i, cell in enumerate(ws_det[1], 1):
        cell.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cell.fill  = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()
    ws_det.row_dimensions[1].height = 30

    for row in summary.rows:
        pct_str = f"{row.utilization_pct:.1f}%" if row.utilization_pct is not None else "—"
        status_label = STATUS_LABELS.get(ReconciliationStatus(row.status), row.status)
        ws_det.append([
            row.patient,
            row.geburtsdatum or "—",
            row.tarifcode    or "—",
            row.tarifziffer  or "—",
            row.valid_from   or "—",
            row.valid_to     or "—",
            row.authorized_minutes,
            row.consumed_minutes,
            row.remaining_minutes,
            pct_str,
            row.service_count,
            row.out_of_range_count,
            status_label,
            row.refusal_reason or "—",
        ])
        data_row = ws_det.max_row
        fill_hex  = STATUS_FILL.get(row.status, "FFFFFF")
        font_hex  = STATUS_FONT_COLOR.get(row.status, "000000")
        status_col = len(headers) - 1   # Status column (second to last)
        reason_col = len(headers)        # Ablehnungsgrund column (last)
        for col in range(1, len(headers) + 1):
            c = ws_det.cell(row=data_row, column=col)
            c.border    = _thin()
            c.alignment = Alignment(vertical="center")
            if col == status_col:
                c.fill = PatternFill("solid", fgColor=fill_hex)
                c.font = Font(name="Calibri", size=10, bold=True, color=font_hex)
            elif col == reason_col:
                c.font = Font(name="Calibri", size=9, color="9C0006" if row.refusal_reason else "595959")
                c.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                c.font = Font(name="Calibri", size=10)

    # Column widths
    col_widths = [28, 14, 12, 12, 12, 12, 16, 16, 18, 13, 18, 14, 22, 36]
    for i, w in enumerate(col_widths, 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws_det.freeze_panes = "A2"

    # ---- Sheet 3: Plan vs. Actual (optional) ----
    if plan_actual:
        overall_fill = {
            OverallStatus.GREEN.value: "C6EFCE",
            OverallStatus.YELLOW.value: "FFEB9C",
            OverallStatus.RED.value: "FFC7CE",
        }
        overall_font = {
            OverallStatus.GREEN.value: "276221",
            OverallStatus.YELLOW.value: "9C5700",
            OverallStatus.RED.value: "9C0006",
        }

        prescription_label = {
            ReconciliationStatus.WITHIN_LIMIT.value: "Gruen",
            ReconciliationStatus.NO_SERVICES.value: "Gruen",
            ReconciliationStatus.NEAR_LIMIT.value: "Gelb",
            ReconciliationStatus.EXCEEDED.value: "Rot",
            ReconciliationStatus.OUT_OF_RANGE.value: "Rot",
            ReconciliationStatus.NO_PRESCRIPTION.value: "Rot",
        }

        ws_pa = wb.create_sheet("Plan vs. Ist")

        ws_pa["A1"] = "Plan-vs.-Ist-Abgleich"
        ws_pa["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F3864")
        ws_pa["A2"] = f"Erstellt am: {plan_actual.generated_at}"
        ws_pa["A2"].font = Font(name="Calibri", size=10, color="595959")
        ws_pa.merge_cells("A1:F1")
        ws_pa.merge_cells("A2:F2")

        ws_pa.append([])
        ws_pa.append(["Kategorie", "Anzahl"])
        for cell in ws_pa[ws_pa.max_row]:
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3864")

        for label, val in [
            ("Gesamt", plan_actual.total_rows),
            ("Gruen", plan_actual.green),
            ("Gelb", plan_actual.yellow),
            ("Rot", plan_actual.red),
            ("OK", plan_actual.ok),
            ("Kleine Abweichung", plan_actual.minor_deviation),
            ("Grosse Abweichung", plan_actual.major_deviation),
            ("Nur verrechnet", plan_actual.billed_not_planned),
            ("Nur geplant", plan_actual.planned_not_billed),
            ("Tarifabweichung", plan_actual.tariff_mismatch),
        ]:
            ws_pa.append([label, val])

        ws_pa.append([])

        pa_headers = [
            "Patient",
            "Datum",
            "Geplante Zeit",
            "Erfasste Zeit",
            "Tarif",
            "Prescription Status",
            "Overall Status",
            "Detailstatus",
            "Geplante Min.",
            "Erbrachte Min.",
            "Abweichung Min.",
            "Hinweis",
        ]

        ws_pa.append(pa_headers)
        for cell in ws_pa[ws_pa.max_row]:
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _thin()
        ws_pa.row_dimensions[ws_pa.max_row].height = 30
        detail_start = ws_pa.max_row

        for pa_row in plan_actual.rows:
            detail_label = PLAN_ACTUAL_LABELS.get(PlanActualStatus(pa_row.status), pa_row.status)
            ws_pa.append([
                pa_row.patient,
                pa_row.service_date or "—",
                pa_row.planned_time or "—",
                pa_row.recorded_time or "—",
                pa_row.tariff or pa_row.tarifziffer_plan or pa_row.tarifziffer_actual or "—",
                prescription_label.get(pa_row.prescription_status, pa_row.prescription_status or "—"),
                pa_row.overall_status or "—",
                detail_label,
                pa_row.planned_minutes if pa_row.planned_minutes is not None else "—",
                pa_row.actual_minutes if pa_row.actual_minutes is not None else "—",
                pa_row.deviation_minutes if pa_row.deviation_minutes is not None else "—",
                pa_row.note or "—",
            ])

            dr = ws_pa.max_row
            row_fill = overall_fill.get(pa_row.overall_status or "", "FFFFFF")
            row_font = overall_font.get(pa_row.overall_status or "", "000000")

            for col in range(1, len(pa_headers) + 1):
                c = ws_pa.cell(row=dr, column=col)
                c.border = _thin()
                c.alignment = Alignment(vertical="center")
                c.font = Font(name="Calibri", size=10)

                if col == 7:  # Overall status
                    c.fill = PatternFill("solid", fgColor=row_fill)
                    c.font = Font(name="Calibri", size=10, bold=True, color=row_font)
                elif col == 8:  # Detail status
                    c.font = Font(name="Calibri", size=10, color="404040")
                elif col == 12:  # Hinweis
                    c.alignment = Alignment(vertical="center", wrap_text=True)

        pa_col_widths = [28, 12, 14, 14, 14, 16, 14, 24, 14, 14, 16, 36]
        for i, w in enumerate(pa_col_widths, 1):
            ws_pa.column_dimensions[get_column_letter(i)].width = w
        ws_pa.freeze_panes = f"A{detail_start + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_csv_report(
    summary: ReconciliationSummary,
    plan_actual: Optional[PlanActualSummary] = None,
) -> bytes:
    """Return CSV bytes (UTF-8 BOM) for easy opening in Excel."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    if plan_actual:
        writer.writerow([
            "Patient",
            "Datum",
            "Geplante Zeit",
            "Erfasste Zeit",
            "Tarif",
            "Prescription Status",
            "Overall Status",
            "Detailstatus",
            "Geplante Min.",
            "Erbrachte Min.",
            "Abweichung Min.",
            "Hinweis",
        ])
        for row in plan_actual.rows:
            detail_label = PLAN_ACTUAL_LABELS.get(PlanActualStatus(row.status), row.status)
            writer.writerow([
                row.patient,
                row.service_date or "",
                row.planned_time or "",
                row.recorded_time or "",
                row.tariff or row.tarifziffer_plan or row.tarifziffer_actual or "",
                row.prescription_status or "",
                row.overall_status or "",
                detail_label,
                row.planned_minutes if row.planned_minutes is not None else "",
                row.actual_minutes if row.actual_minutes is not None else "",
                row.deviation_minutes if row.deviation_minutes is not None else "",
                row.note or "",
            ])
    else:
        writer.writerow([
            "Patient",
            "Geburtsdatum",
            "Tarifcode",
            "Tarifziffer",
            "Gueltig von",
            "Gueltig bis",
            "Bewilligte Min.",
            "Erbrachte Min.",
            "Verbleibend Min.",
            "Auslastung %",
            "Leistungen",
            "Status",
            "Ablehnungsgrund",
        ])
        for row in summary.rows:
            writer.writerow([
                row.patient,
                row.geburtsdatum or "",
                row.tarifcode or "",
                row.tarifziffer or "",
                row.valid_from or "",
                row.valid_to or "",
                row.authorized_minutes,
                row.consumed_minutes,
                row.remaining_minutes,
                row.utilization_pct if row.utilization_pct is not None else "",
                row.service_count,
                row.status,
                row.refusal_reason or "",
            ])

    return output.getvalue().encode("utf-8-sig")
